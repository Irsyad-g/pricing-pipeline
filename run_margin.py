import pandas as pd
import numpy as np
import json
import re
import argparse
from pathlib import Path
from datetime import datetime
from sqlalchemy import text
from config.database import get_engine
from config.commission import get_commission
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

CNY_RATE          = 2650

RAW_DIR           = Path("data/raw")
MCC_MAP_PATH      = Path("data/mappings/mcc_map.json")
COUNTRY_RATE_PATH = Path("data/mappings/country_rate.json")
OUTPUT_DIR        = Path("data/output")
MANUAL_INPUT_PATH = OUTPUT_DIR / "manual_input.xlsx"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

with open(MCC_MAP_PATH)      as f: MCC_MAP      = json.load(f)
with open(COUNTRY_RATE_PATH) as f: COUNTRY_RATE = json.load(f)

GLOBAL_RATE = np.median(list(COUNTRY_RATE.values()))

BACKLOG_COLS = ["Channel", "Invoice", "ICCID", "SKU", "Product_Name",
                "Product_Type", "Order_Date", "Month", "Harga_Jual"]

# ── SKU normalization (single place, reusable) ────────────────
SKU_PREFIX_RULES = [
    (r"^TU-GK-",   "GK-"),
    (r"^GM-GK-",   "GK-"),
    (r"^MM-GK-",   "GK-"),
]

SKU_MARKETPLACE_RULES = SKU_PREFIX_RULES + [
    (r"^GK-SEA4-",  "GK-SEA5-"),
    (r"^GK-SEA6-",  "GK-SEA5-"),
    (r"^GK-JPN-",   "GK-JPNPLUS-"),
    (r"^WG-SAU-",   "GK-SAU-"),
    (r"^WG-SGMY-",  "GK-SGMY-"),
    (r"^WG-JPN-",   "GK-JPNPLUS-"),
    (r"^WG-JPNM-",  "GK-JPNPLUS-"),
    (r"^WG-SEA5-",  "GK-SEA5-"),
    (r"^WG-DTAC-",  "GK-THA-"),
    (r"^SM-USA-",   "GK-USA-"),
    (r"^GM-",       "GK-"),
]

VALID_SKU_PREFIXES = ("GK-", "GM-GK-", "TU-GK", "MM-GK", "WG-", "SM-")


def normalize_sku(sku, rules=None):
    """Apply SKU normalization rules."""
    if not sku or not isinstance(sku, str):
        return sku
    s = sku.strip().upper()
    for pattern, replacement in (rules or SKU_PREFIX_RULES):
        s = re.sub(pattern, replacement, s)
    return s


def extract_mcc(country_str):
    m = re.match(r"(\d+)", str(country_str).strip())
    return m.group(1) if m else None


def get_rate(country_str):
    mcc          = extract_mcc(country_str)
    country_name = MCC_MAP.get(str(mcc)) if mcc else None
    rate         = COUNTRY_RATE.get(country_name) if country_name else None
    return rate if rate else GLOBAL_RATE


def parse_args():
    parser = argparse.ArgumentParser(description="Generate margin report")
    parser.add_argument("--start-date", type=str, default=None,
                        help="Start date filter (YYYY-MM-DD). Filter output only.")
    parser.add_argument("--end-date", type=str, default=None,
                        help="End date filter (YYYY-MM-DD). Filter output only.")
    parser.add_argument("--output", type=str, default=None,
                        help="Custom output filename (default: margin_report.xlsx or with date range)")
    return parser.parse_args()


def get_output_path(args):
    if args.output:
        return OUTPUT_DIR / args.output
    if args.start_date and args.end_date:
        return OUTPUT_DIR / f"margin_report_{args.start_date}_to_{args.end_date}.xlsx"
    if args.start_date:
        return OUTPUT_DIR / f"margin_report_from_{args.start_date}.xlsx"
    if args.end_date:
        return OUTPUT_DIR / f"margin_report_until_{args.end_date}.xlsx"
    return OUTPUT_DIR / "margin_report.xlsx"


def load_orders():
    files = list(RAW_DIR.glob("orders_export*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Tidak ada file orders_export*.xlsx di {RAW_DIR}")

    print(f"  Ditemukan {len(files)} file orders_export:")
    dfs = []
    for f in files:
        print(f"  - {f.name}")
        df = pd.read_excel(f, header=0, dtype=str)
        temp = pd.DataFrame({
            "Channel":      df.iloc[:, 0],
            "Status":       df.iloc[:, 2],
            "Invoice":      df.iloc[:, 3],
            "Order_Date":   df.iloc[:, 7],
            "Product_Name": df.iloc[:, 14],
            "SKU_Original": df.iloc[:, 13],
            "Product_Type": df.iloc[:, 11],
            "Harga_Jual":   df.iloc[:, 17],
            "ICCID":        df.iloc[:, 19],
        })
        dfs.append(temp)

    orders = pd.concat(dfs, ignore_index=True)
    orders["Status_clean"] = orders["Status"].astype(str).str.strip().str.upper()
    orders["status_priority"] = np.where(orders["Status_clean"] == "DONE", 1, 2)

    # Shopify: capture original item count per Invoice BEFORE ICCID deduplication.
    # Duplicate ICCIDs in the same order reduce post-dedup count, causing wrong per-unit price.
    _shopify_mask = orders["Channel"].str.contains("Shopify", case=False, na=False)
    if _shopify_mask.any():
        _shopify_counts = (
            orders[_shopify_mask]
            .groupby("Invoice")
            .size()
            .reset_index(name="_shopify_item_count")
        )
        orders = orders.merge(_shopify_counts, on="Invoice", how="left")

    orders = (
        orders
        .sort_values(["ICCID", "status_priority", "Order_Date"],
                      ascending=[True, True, False])
        .drop_duplicates(subset=["ICCID"], keep="first")
    )
    orders["Order_Date"]   = pd.to_datetime(orders["Order_Date"], errors="coerce")
    orders["Month"]        = orders["Order_Date"].dt.strftime("%Y-%b")
    orders["Channel"]      = orders["Channel"].str.strip()
    orders["Invoice"]      = orders["Invoice"].str.strip().str.lstrip("'")
    orders["Product_Name"] = orders["Product_Name"].str.strip()
    orders["SKU_Original"] = orders["SKU_Original"].str.strip().str.upper()
    orders["Product_Type"] = orders["Product_Type"].str.strip()
    orders["ICCID"]        = orders["ICCID"].str.strip().str.lstrip("'")
    orders["Harga_Jual"]   = pd.to_numeric(orders["Harga_Jual"], errors="coerce").fillna(0)

    before = len(orders)
    orders = orders.drop_duplicates()
    print(f"  Duplikat dihapus: {before - len(orders)} baris")

    # filter ICCID valid
    orders = orders[orders["ICCID"].str.startswith("898", na=False)]
    orders = orders[orders["SKU_Original"].notna() & (orders["SKU_Original"] != "")]

    # expand SKU prefix filter — include WG- dan SM- yang akan di-normalize
    orders = orders[orders["SKU_Original"].str.startswith(VALID_SKU_PREFIXES, na=False)]

    orders["Product_Type"] = np.where(
        orders["SKU_Original"].str.startswith("GM-GK-", na=False),
        "Simcard", orders["Product_Type"]
    )

    # normalize SKU (single place)
    orders["SKU"] = orders["SKU_Original"].apply(
        lambda s: normalize_sku(s, SKU_PREFIX_RULES)
    )

    return orders


def load_legacy_orders():
    legacy_path = RAW_DIR / "orders_legacy.xlsx"
    if not legacy_path.exists():
        print("  Tidak ada file legacy orders")
        return pd.DataFrame(columns=["Invoice", "ICCID", "SKU", "Order_Date"])

    df = pd.read_excel(legacy_path, dtype=str)
    df["Invoice"]    = df["Invoice"].str.strip().str.lstrip("'")
    df["ICCID"]      = df["ICCID"].str.strip().str.lstrip("'")
    df["SKU"]        = df["SKU"].str.strip().str.upper().apply(
        lambda s: normalize_sku(s, SKU_PREFIX_RULES)
    )
    df["Order_Date"] = pd.to_datetime(df["Order_Date"], errors="coerce")
    df["Month"]      = df["Order_Date"].dt.strftime("%Y-%b")
    df = df.drop_duplicates(subset=["Invoice", "ICCID"])
    print(f"  Legacy orders: {len(df)} baris")
    return df


def load_shopee_revenue():
    files = list(RAW_DIR.glob("Order.all*.xlsx"))
    if not files:
        print("  Tidak ada file Shopee (Order.all*.xlsx)")
        return pd.DataFrame(columns=["Invoice", "Pendapatan_Shopee", "Shopee_Status"])

    print(f"  Ditemukan {len(files)} file Shopee:")
    dfs = []
    for f in files:
        print(f"  - {f.name}")
        df = pd.read_excel(f, header=0, dtype=str)
        dfs.append(pd.DataFrame({
            "Invoice":           df.iloc[:, 0],
            "SKU":               df.iloc[:, 14],
            "Pendapatan_Shopee": df.iloc[:, 17],
            "Shopee_Status":     df.iloc[:, 3],
        }))

    shopee = pd.concat(dfs, ignore_index=True)
    shopee["Invoice"] = shopee["Invoice"].str.strip().str.lstrip("'")
    shopee["SKU"] = shopee["SKU"].str.strip().str.upper().apply(
        lambda s: normalize_sku(s, SKU_MARKETPLACE_RULES)
    )
    shopee["Shopee_Status"]     = shopee["Shopee_Status"].str.strip()
    shopee["Pendapatan_Shopee"] = (
        shopee["Pendapatan_Shopee"]
        .str.replace(".", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
        .fillna(0)
    )
    shopee = shopee.drop_duplicates(subset=["Invoice", "SKU"])
    need_manual = (
        (shopee["Pendapatan_Shopee"] == 0) |
        (shopee["Shopee_Status"].str.contains("Permintaan Disetujui", na=False))
    ).sum()
    print(f"  Shopee: {len(shopee)} invoice ({need_manual} perlu input manual)")
    return shopee


def load_manual_input():
    if not MANUAL_INPUT_PATH.exists():
        return pd.DataFrame(columns=["Invoice", "Harga_Final"])

    df = pd.read_excel(MANUAL_INPUT_PATH, dtype=str)
    df["Invoice"]     = df["Invoice"].str.strip().str.lstrip("'")
    df["Harga_Final"] = pd.to_numeric(df["Harga_Final"], errors="coerce")
    filled = df["Harga_Final"].notna().sum()
    print(f"  Manual input: {filled} invoice sudah diisi dari {len(df)} total")
    return df[["Invoice", "Harga_Final"]]


def generate_manual_input(orders, shopee):
    is_shopee = orders["Channel"].str.contains("Shopee", case=False, na=False)
    mask = (
        (orders["Pendapatan_Shopee"].fillna(0) == 0) |
        (orders["Shopee_Status"].str.contains("Permintaan Disetujui", na=False))
    )
    need_manual = orders[is_shopee & mask].copy()
    need_manual = need_manual.drop_duplicates(subset=["Invoice", "ICCID"], keep="first")

    if need_manual.empty:
        print("  Tidak ada invoice yang perlu input manual")
        return

    existing = pd.DataFrame(columns=["Invoice", "ICCID", "Harga_Final"])
    if MANUAL_INPUT_PATH.exists():
        existing = pd.read_excel(MANUAL_INPUT_PATH, dtype=str)
        existing["Invoice"]     = existing["Invoice"].str.strip()
        existing["ICCID"]       = existing["ICCID"].str.strip()
        existing["Harga_Final"] = pd.to_numeric(existing["Harga_Final"], errors="coerce")

    if not existing.empty:
        existing_keys = set(zip(existing["Invoice"], existing["ICCID"]))
        need_manual = need_manual[
            ~need_manual.apply(lambda r: (r["Invoice"], r["ICCID"]) in existing_keys, axis=1)
        ]

    if need_manual.empty:
        print("  Tidak ada invoice baru yang perlu input manual")
        return

    output_new = need_manual[["Channel", "Invoice", "ICCID", "SKU",
                               "Product_Name", "Order_Date",
                               "Pendapatan_Shopee", "Shopee_Status"]].copy()
    output_new["Harga_Final"] = None
    output_new["Catatan"] = output_new["Shopee_Status"].apply(
        lambda s: "Isi Harga_Final — Permintaan Disetujui (refund)"
                  if "Permintaan Disetujui" in str(s)
                  else "Isi Harga_Final — harga tidak terbaca"
    )
    output_new.insert(output_new.columns.get_loc("Pendapatan_Shopee") + 1,
                      "Harga_Final", output_new.pop("Harga_Final"))

    output = pd.concat([existing, output_new], ignore_index=True)

    with pd.ExcelWriter(MANUAL_INPUT_PATH, engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="Input Manual", index=False)
        ws = writer.sheets["Input Manual"]

        hf_col      = output.columns.get_loc("Harga_Final") + 1
        FILL_YELLOW = PatternFill("solid", fgColor="FFFF99")
        FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
        FONT_HEADER = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
        for row in ws.iter_rows(min_row=2):
            row[hf_col - 1].fill = FILL_YELLOW
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    print(f"  Manual input file: {MANUAL_INPUT_PATH} ({len(output_new)} baru, {len(existing)} existing)")


def load_tokopedia_revenue():
    files = list(RAW_DIR.glob("Semua pesanan*.xlsx"))
    if not files:
        print("  Tidak ada file Tokopedia (Semua pesanan*.xlsx)")
        return pd.DataFrame(columns=["Invoice", "Pendapatan_Tokped"])

    print(f"  Ditemukan {len(files)} file Tokopedia:")
    dfs = []
    for f in files:
        print(f"  - {f.name}")
        df = pd.read_excel(f, header=0, dtype=str)
        dfs.append(pd.DataFrame({
            "Invoice":           df.iloc[:, 0],
            "SKU":               df.iloc[:, 6],
            "Pendapatan_Tokped": df.iloc[:, 11],
        }))

    tokped = pd.concat(dfs, ignore_index=True)
    tokped["Invoice"] = tokped["Invoice"].str.strip().str.lstrip("'")
    tokped["SKU"] = tokped["SKU"].str.strip().str.upper().apply(
        lambda s: normalize_sku(s, SKU_MARKETPLACE_RULES)
    )
    tokped["Pendapatan_Tokped"] = pd.to_numeric(tokped["Pendapatan_Tokped"], errors="coerce").fillna(0)
    tokped = tokped[tokped["Pendapatan_Tokped"] > 0].drop_duplicates(subset=["Invoice", "SKU"])
    print(f"  Tokopedia: {len(tokped)} baris")
    return tokped


def load_db_data(engine):
    with engine.connect() as conn:
        country_usage = pd.read_sql(text("""
            SELECT iccid, country, SUM(country_usage_mb) as usage_mb
            FROM processed.country_usage
            GROUP BY iccid, country
        """), conn)

        final = pd.read_sql(text("""
            SELECT iccid, package, start_date, end_date,
                   total_usage_mb, total_quota_mb,
                   real_cost_cny, real_cost_idr
            FROM processed.final_output
        """), conn)

        # load semua status (active + expired) untuk coverage lengkap
        all_sub = pd.read_sql(text("""
            SELECT iccid, package, status, start_date
            FROM raw.subscription
        """), conn)

    # strip iccid — nilai DB bisa punya whitespace tersembunyi
    country_usage["iccid"] = country_usage["iccid"].astype(str).str.strip()

    if country_usage.empty:
        # processed.country_usage belum diisi upstream pipeline (db_exporter tidak menulis tabel ini)
        # fallback ke raw.daily_usage yang sudah ada di DB dengan struktur MCC yang sama
        print("  [WARNING] processed.country_usage kosong — upstream pipeline belum populate tabel ini.")
        print("  [FALLBACK] Menggunakan raw.daily_usage sebagai sumber country breakdown.")
        with engine.connect() as conn:
            country_usage = pd.read_sql(text("""
                SELECT iccid, area AS country, SUM(usage_mb) AS usage_mb
                FROM raw.daily_usage
                GROUP BY iccid, area
            """), conn)
        country_usage["iccid"] = country_usage["iccid"].astype(str).str.strip()
        if country_usage.empty:
            print("  [WARNING] raw.daily_usage juga kosong. Negara_Detail akan kosong.")
    else:
        print(f"  country_usage: {len(country_usage)} rows dari processed.country_usage")

    # active_sub: non-expired only — dipakai untuk klasifikasi ACTIVE di matching
    active_sub = all_sub[all_sub["status"].str.lower() != "expired"].copy().reset_index(drop=True)

    return country_usage, final, active_sub


def apply_date_filter(df, start_date=None, end_date=None):
    """Filter DataFrame by Order_Date range. Returns filtered copy."""
    filtered = df.copy()
    if start_date:
        sd = pd.to_datetime(start_date)
        filtered = filtered[filtered["Order_Date"] >= sd]
        print(f"  Date filter: >= {start_date} ({len(df) - len(filtered)} excluded)")
    if end_date:
        ed = pd.to_datetime(end_date)
        filtered = filtered[filtered["Order_Date"] <= ed]
        remaining = len(df) - len(filtered)
        print(f"  Date filter: <= {end_date}")
    if start_date or end_date:
        print(f"  After date filter: {len(filtered)} dari {len(df)} baris")
    return filtered


def extract_fallback_candidates(diagnostics_df):
    """
    Parse diagnostics SKU_Mismatch_Detail untuk temukan country code mismatch.
    Returns DataFrame: Order_Country, DB_Country, Count_ICCID, Example_SKUs
    Hanya ambil kasus ICCID ada di DB (bukan missing ICCID) — pure SKU country mismatch.
    """
    if diagnostics_df.empty:
        return pd.DataFrame()

    db_exists = diagnostics_df[
        diagnostics_df["DB_ICCID_Exists"].astype(str).str.lower() == "true"
    ].copy()
    if db_exists.empty:
        return pd.DataFrame()

    country_pattern = re.compile(r"country: order=(\w+) vs db=(\w+)")
    rows = []
    for _, row in db_exists.iterrows():
        detail = str(row.get("SKU_Mismatch_Detail", ""))
        for m in country_pattern.finditer(detail):
            order_c, db_c = m.group(1), m.group(2)
            if order_c != db_c:
                rows.append({
                    "Order_SKU":     row["Order_SKU"],
                    "Order_Country": order_c,
                    "DB_Country":    db_c,
                    "ICCID":         row["ICCID"],
                })

    if not rows:
        return pd.DataFrame()

    cand_df = pd.DataFrame(rows)
    summary = (
        cand_df.groupby(["Order_Country", "DB_Country"])
        .agg(
            Count_ICCID  = ("ICCID",     "nunique"),
            Example_SKUs = ("Order_SKU", lambda x: " | ".join(sorted(x.unique())[:3]))
        )
        .reset_index()
        .sort_values("Count_ICCID", ascending=False)
        .reset_index(drop=True)
    )
    summary["Already_In_Fallback"] = summary["Order_Country"].isin(
        list(__import__("json").load(open(Path("data/mappings/sku_fallback.json"))).keys())
    )
    return summary


def style_sheet(ws, pct_cols=(), idr_cols=(), cny_cols=(), status_col=None):
    FILL_RUGI   = PatternFill("solid", fgColor="FFCCCC")
    FILL_BAGUS  = PatternFill("solid", fgColor="CCFFCC")
    FILL_NORMAL = PatternFill("solid", fgColor="FFF9CC")
    FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    ALIGN_CTR   = Alignment(horizontal="center", vertical="center")
    STATUS_FILL = {"RUGI": FILL_RUGI, "BAGUS": FILL_BAGUS, "NORMAL": FILL_NORMAL}
    IDR_FMT = '"Rp"#,##0'
    CNY_FMT = '"¥"#,##0.00'
    PCT_FMT = '0.00%'

    for cell in ws[1]:
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER; cell.alignment = ALIGN_CTR
    headers = {cell.value: cell.column for cell in ws[1]}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = ws.cell(1, cell.column).value
            if col_name in pct_cols:
                cell.number_format = PCT_FMT; cell.alignment = ALIGN_CTR
            elif col_name in idr_cols:
                cell.number_format = IDR_FMT
            elif col_name in cny_cols:
                cell.number_format = CNY_FMT
        if status_col and status_col in headers:
            stat = ws.cell(row[0].row, headers[status_col]).value
            fill = STATUS_FILL.get(stat)
            if fill:
                for cell in row: cell.fill = fill
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _write_settings_sheet(wb, rate):
    ws = wb.create_sheet("Settings")
    FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
    FONT_HEADER = Font(color="FFFFFF", bold=True)
    FILL_INPUT  = PatternFill("solid", fgColor="FFFF99")
    FONT_BOLD   = Font(bold=True)

    for ci, h in enumerate(["Parameter", "Nilai"], 1):
        c = ws.cell(1, ci, h)
        c.fill = FILL_HEADER; c.font = FONT_HEADER

    settings = [
        ("CNY_TO_IDR Rate (ubah di sini jika kurs berubah)", rate),
        ("Margin RUGI threshold (<)", 0.0),
        ("Margin BAGUS threshold (>=)", 0.30),
    ]
    for ri, (name, val) in enumerate(settings, 2):
        ws.cell(ri, 1, name).font = FONT_BOLD
        c = ws.cell(ri, 2, val)
        c.fill = FILL_INPUT
        if ri > 2:
            c.number_format = "0.00%"
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 15


def _write_summary_sku_formula_sheet(ws, df, last_detail_row):
    """Summary per SKU with SUMIF formulas referencing Detail per ICCID sheet."""
    FILL_RAW    = PatternFill("solid", fgColor="DDEEFF")
    FILL_FORM   = PatternFill("solid", fgColor="FFFACD")
    FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    ALIGN_CTR   = Alignment(horizontal="center", vertical="center")
    FILL_RUGI   = PatternFill("solid", fgColor="FFCCCC")
    FILL_BAGUS  = PatternFill("solid", fgColor="CCFFCC")
    FILL_NORMAL = PatternFill("solid", fgColor="FFF9CC")
    IDR = '"Rp"#,##0'
    PCT = "0.00%"

    # Column letters in 'Detail per ICCID' (matches _write_detail_formula_sheet layout)
    DET = "'Detail per ICCID'"
    D_SKU  = "E";  D_INV  = "C";  D_HARGA = "H"
    D_KOM  = "M";  D_NET  = "N";  D_CIDR  = "O";  D_MARGIN = "S"
    ldr = last_detail_row

    headers = ["SKU", "Total_Invoice", "Total_ICCID", "Total_Pendapatan",
               "Total_Komisi", "Total_Net_Revenue", "Total_Cost_IDR",
               "Total_Margin_IDR", "Avg_Margin_Pct", "Status"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER; cell.alignment = ALIGN_CTR

    # Sort SKUs by margin ascending (same order as Python groupby)
    sku_list = df.groupby("SKU")["Margin_IDR"].sum().sort_values().index.tolist()

    # Column letter shortcuts for this sheet
    A = "A"; F = "F"; H = "H"; I = "I"

    for ri, sku in enumerate(sku_list, 2):
        # A: SKU (raw)
        ws.cell(ri, 1, sku).fill = FILL_RAW

        # B: Total_Invoice — unique invoice count for this SKU via SUMPRODUCT
        ws.cell(ri, 2,
            f"=SUMPRODUCT(({DET}!${D_SKU}$2:${D_SKU}${ldr}={A}{ri})"
            f"*IFERROR(1/COUNTIFS({DET}!${D_INV}$2:${D_INV}${ldr},{DET}!${D_INV}$2:${D_INV}${ldr}"
            f",{DET}!${D_SKU}$2:${D_SKU}${ldr},{A}{ri}),0))"
        ).fill = FILL_FORM

        # C: Total_ICCID
        ws.cell(ri, 3,
            f"=COUNTIF({DET}!${D_SKU}$2:${D_SKU}${ldr},{A}{ri})"
        ).fill = FILL_FORM

        def sf(col_letter):
            return f"=SUMIF({DET}!${D_SKU}$2:${D_SKU}${ldr},{A}{ri},{DET}!${col_letter}$2:${col_letter}${ldr})"

        c = ws.cell(ri, 4, sf(D_HARGA));   c.fill = FILL_FORM; c.number_format = IDR  # Total_Pendapatan
        c = ws.cell(ri, 5, sf(D_KOM));     c.fill = FILL_FORM; c.number_format = IDR  # Total_Komisi
        c = ws.cell(ri, 6, sf(D_NET));     c.fill = FILL_FORM; c.number_format = IDR  # Total_Net_Revenue
        c = ws.cell(ri, 7, sf(D_CIDR));    c.fill = FILL_FORM; c.number_format = IDR  # Total_Cost_IDR
        c = ws.cell(ri, 8, sf(D_MARGIN));  c.fill = FILL_FORM; c.number_format = IDR  # Total_Margin_IDR

        # I: Avg_Margin_Pct = Total_Margin_IDR / Total_Net_Revenue
        c = ws.cell(ri, 9, f"=IF({F}{ri}>0,{H}{ri}/{F}{ri},0)")
        c.fill = FILL_FORM; c.number_format = PCT

        # J: Status
        ws.cell(ri, 10,
            f'=IF({I}{ri}<0,"RUGI",IF({I}{ri}>=0.3,"BAGUS","NORMAL"))'
        ).fill = FILL_FORM

    last_row = len(sku_list) + 1
    ws.conditional_formatting.add(f"A2:J{last_row}",
        FormulaRule(formula=[f'$J2="RUGI"'],   fill=FILL_RUGI))
    ws.conditional_formatting.add(f"A2:J{last_row}",
        FormulaRule(formula=[f'$J2="BAGUS"'],  fill=FILL_BAGUS))
    ws.conditional_formatting.add(f"A2:J{last_row}",
        FormulaRule(formula=[f'$J2="NORMAL"'], fill=FILL_NORMAL))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _write_detail_formula_sheet(ws, df):
    FILL_RAW    = PatternFill("solid", fgColor="DDEEFF")
    FILL_FORM   = PatternFill("solid", fgColor="FFFACD")
    FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
    FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
    ALIGN_CTR   = Alignment(horizontal="center", vertical="center")
    FILL_RUGI   = PatternFill("solid", fgColor="FFCCCC")
    FILL_BAGUS  = PatternFill("solid", fgColor="CCFFCC")
    FILL_NORMAL = PatternFill("solid", fgColor="FFF9CC")

    IDR = '"Rp"#,##0'
    CNY = '"¥"#,##0.00'
    PCT = "0.00%"

    # Column positions (1-based)
    C = {
        "Channel": 1, "Product_Name": 2, "Invoice": 3, "ICCID": 4,
        "SKU": 5, "Product_Type": 6, "Month": 7,
        "Harga_Jual": 8, "Komisi_Pct": 9, "Real_Cost_CNY": 10,
        "Total_Usage_MB": 11, "Total_Quota_MB": 12,
        # formula columns
        "Komisi_IDR": 13, "Net_Revenue": 14, "Real_Cost_IDR": 15,
        "Total_Usage_Display": 16, "Ratio": 17, "Total_Quota_Display": 18,
        "Margin_IDR": 19, "Margin_Pct": 20, "Status": 21,
        # info
        "Match_Method": 22, "Negara_Detail": 23,
    }

    def cl(key): return get_column_letter(C[key])

    headers = list(C.keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = FILL_HEADER; cell.font = FONT_HEADER; cell.alignment = ALIGN_CTR

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        def w(key, val, fmt=None, fill=None):
            cell = ws.cell(ri, C[key], val)
            if fmt:  cell.number_format = fmt
            if fill: cell.fill = fill

        val_or_none = lambda v: v if pd.notna(v) else None
        val_or_zero = lambda v: float(v) if pd.notna(v) else 0.0

        w("Channel",       val_or_none(row.get("Channel")),      fill=FILL_RAW)
        w("Product_Name",  val_or_none(row.get("Product_Name")), fill=FILL_RAW)
        w("Invoice",       val_or_none(row.get("Invoice")),      fill=FILL_RAW)
        w("ICCID",         val_or_none(row.get("ICCID")),        fill=FILL_RAW)
        w("SKU",           val_or_none(row.get("SKU")),          fill=FILL_RAW)
        w("Product_Type",  val_or_none(row.get("Product_Type")), fill=FILL_RAW)
        w("Month",         val_or_none(row.get("Month")),        fill=FILL_RAW)
        w("Harga_Jual",    val_or_zero(row.get("Harga_Jual")),   fmt=IDR, fill=FILL_RAW)
        w("Komisi_Pct",    val_or_zero(row.get("Komisi_Pct")),   fmt=PCT, fill=FILL_RAW)
        w("Real_Cost_CNY", val_or_zero(row.get("Real_Cost_CNY")), fmt=CNY, fill=FILL_RAW)
        w("Total_Usage_MB", val_or_zero(row.get("Total_Usage_MB")), fill=FILL_RAW)
        w("Total_Quota_MB", val_or_zero(row.get("Total_Quota_MB")), fill=FILL_RAW)

        H = cl("Harga_Jual"); I = cl("Komisi_Pct"); J = cl("Real_Cost_CNY")
        K = cl("Total_Usage_MB"); L = cl("Total_Quota_MB")
        M = cl("Komisi_IDR");    N = cl("Net_Revenue"); O = cl("Real_Cost_IDR")
        S = cl("Margin_IDR");    T = cl("Margin_Pct")

        w("Komisi_IDR",          f"={H}{ri}*{I}{ri}",                           fmt=IDR, fill=FILL_FORM)
        w("Net_Revenue",         f"={H}{ri}-{M}{ri}",                           fmt=IDR, fill=FILL_FORM)
        w("Real_Cost_IDR",       f"={J}{ri}*Settings!$B$2",                     fmt=IDR, fill=FILL_FORM)
        w("Total_Usage_Display", f'=IF({K}{ri}>=1024,TEXT({K}{ri}/1024,"0.00")&" GB",TEXT({K}{ri},"0.00")&" MB")', fill=FILL_FORM)
        w("Ratio",               f"=IF({L}{ri}>0,{K}{ri}/{L}{ri},0)",           fmt=PCT, fill=FILL_FORM)
        w("Total_Quota_Display", f'=IF({L}{ri}>=1024,TEXT({L}{ri}/1024,"0.00")&" GB",TEXT({L}{ri},"0.00")&" MB")', fill=FILL_FORM)
        w("Margin_IDR",          f"={N}{ri}-{O}{ri}",                           fmt=IDR, fill=FILL_FORM)
        w("Margin_Pct",          f"=IF({N}{ri}>0,{S}{ri}/{N}{ri},0)",           fmt=PCT, fill=FILL_FORM)
        w("Status",              f'=IF({T}{ri}<0,"RUGI",IF({T}{ri}>=0.3,"BAGUS","NORMAL"))', fill=FILL_FORM)

        w("Match_Method",  val_or_none(row.get("Match_Method", "SKU_MATCH")))
        w("Negara_Detail", val_or_none(row.get("Negara_Detail", "UNKNOWN")))

    last_row = len(df) + 1
    sc = cl("Status"); last_col = cl("Negara_Detail")
    data_range = f"A2:{last_col}{last_row}"
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f"${sc}2=\"RUGI\""],   fill=FILL_RUGI))
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f"${sc}2=\"BAGUS\""],  fill=FILL_BAGUS))
    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=[f"${sc}2=\"NORMAL\""], fill=FILL_NORMAL))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


if __name__ == "__main__":
    args = parse_args()
    OUTPUT = get_output_path(args)

    print("Loading orders...")
    orders = load_orders()
    backlog = pd.DataFrame(columns=BACKLOG_COLS + ["Backlog_Reason"])

    print("\nLoading Shopee revenue...")
    shopee = load_shopee_revenue()

    print("\nLoading Tokopedia revenue...")
    tokped = load_tokopedia_revenue()

    orders = orders.merge(shopee, on=["Invoice", "SKU"], how="left")
    orders = orders.merge(tokped, on=["Invoice", "SKU"], how="left")

    # load legacy
    print("\nLoading legacy orders...")
    legacy = load_legacy_orders()

    if not legacy.empty:
        legacy = legacy.merge(shopee[["Invoice", "SKU", "Pendapatan_Shopee", "Shopee_Status"]],
                              on=["Invoice", "SKU"], how="left")
        legacy = legacy.merge(tokped[["Invoice", "SKU", "Pendapatan_Tokped"]],
                              on=["Invoice", "SKU"], how="left")

        for col in ["Channel", "Order_Date", "Product_Name", "Product_Type", "Month"]:
            if col not in legacy.columns:
                legacy[col] = "Legacy"
        legacy["Order_Date"] = pd.to_datetime(legacy["Order_Date"], errors="coerce")
        legacy["Month"]      = legacy["Order_Date"].dt.strftime("%Y-%b").fillna("2026-Feb")

        is_shopee_leg = legacy["Pendapatan_Shopee"].fillna(0) > 0
        is_tokped_leg = legacy["Pendapatan_Tokped"].fillna(0) > 0
        legacy["Harga_Jual"] = np.select(
            [is_shopee_leg, is_tokped_leg],
            [legacy["Pendapatan_Shopee"], legacy["Pendapatan_Tokped"]],
            default=0
        )

        orders = pd.concat([orders, legacy], ignore_index=True)
        print(f"  Legacy digabung: {len(legacy)} baris")

    is_shopee = orders["Channel"].str.contains("Shopee",    case=False, na=False)
    is_tokped = orders["Channel"].str.contains("Tokopedia", case=False, na=False)
    orders["Harga_Jual"] = np.select(
        [is_shopee, is_tokped],
        [orders["Pendapatan_Shopee"].fillna(0),
         orders["Pendapatan_Tokped"].fillna(0)],
        default=orders["Harga_Jual"]
    )

    # Shopify: Harga_Jual di export = total order, bukan per-unit.
    # Divide by original item count (before ICCID dedup) to avoid undercounting when
    # duplicate ICCIDs exist in the same order.
    is_shopify = orders["Channel"].str.contains("Shopify", case=False, na=False)
    if is_shopify.any():
        shopify_idx   = orders[is_shopify].index
        grp           = orders.loc[shopify_idx].groupby(["Invoice", "SKU"])["Harga_Jual"]
        max_per_grp   = grp.transform("max")
        if "_shopify_item_count" in orders.columns:
            count_per_grp = orders.loc[shopify_idx, "_shopify_item_count"].fillna(
                grp.transform("count")
            )
        else:
            count_per_grp = grp.transform("count")
        orders.loc[shopify_idx, "Harga_Jual"] = (max_per_grp / count_per_grp).round(0)
        print(f"  Shopify: per-unit price dihitung untuk {is_shopify.sum()} ICCID "
              f"({(orders.loc[shopify_idx, 'Harga_Jual'] == 0).sum()} masih 0 → backlog)")

    print("\nGenerating manual input file...")
    generate_manual_input(orders, shopee)

    print("\nLoading manual input...")
    has_manual = pd.Series(False, index=orders.index)

    manual = load_manual_input()
    if not manual.empty:
        orders = orders.merge(manual, on="Invoice", how="left")
        orders = orders.reset_index(drop=True)
        has_manual = orders["Harga_Final"].notna()
        orders.loc[has_manual, "Harga_Jual"] = orders.loc[has_manual, "Harga_Final"]
        orders = orders.drop(columns=["Harga_Final"])
        print(f"  Manual input diterapkan: {has_manual.sum()} invoice")

    has_manual = has_manual.reset_index(drop=True) if not manual.empty else pd.Series(False, index=orders.index)

    bl_refund = orders[has_manual & (orders["Harga_Jual"] == 0)][BACKLOG_COLS].copy()
    bl_refund["Backlog_Reason"] = "FULL_REFUND"

    bl_zero = orders[~has_manual & (orders["Harga_Jual"] == 0)][BACKLOG_COLS].copy()
    bl_zero["Backlog_Reason"] = "HARGA_0_SETELAH_MERGE"
    backlog = pd.concat([bl_refund, bl_zero], ignore_index=True)

    print("\nLoading data dari DB...")
    engine                           = get_engine()
    country_usage, final, active_sub = load_db_data(engine)

    print("\nMatching orders ke subscriptions...")
    from processors.cost_calculator import match_orders_to_subscriptions
    orders, diagnostics_df = match_orders_to_subscriptions(orders, final, active_sub)

    # pisahkan yang tidak ter-match → backlog
    no_match = orders[orders["real_cost_idr"].isna()].copy()
    orders   = orders[orders["real_cost_idr"].notna() & (orders["real_cost_idr"] > 0)].copy()
    orders   = orders[orders["Harga_Jual"] > 0].copy()

    if not no_match.empty:
        bl_nomatch = no_match[BACKLOG_COLS + ["Match_Method"]].copy()
        bl_nomatch = bl_nomatch.rename(columns={"Match_Method": "Backlog_Reason"})
        backlog    = pd.concat([backlog, bl_nomatch], ignore_index=True)

        backlog_db = no_match[["Channel", "Invoice", "ICCID", "SKU", "Product_Name",
                                "Product_Type", "Order_Date", "Harga_Jual", "Match_Method"]].copy()
        backlog_db.columns = ["channel", "invoice", "iccid", "sku", "product_name",
                               "product_type", "order_date", "harga_jual", "reason"]
        backlog_db["resolved"] = False

        with engine.begin() as conn:
            for _, row in backlog_db.iterrows():
                exists = conn.execute(text("""
                    SELECT 1 FROM processed.margin_backlog
                    WHERE invoice = :inv AND iccid = :iccid
                """), {"inv": row["invoice"], "iccid": row["iccid"]}).fetchone()
                if not exists:
                    conn.execute(text("""
                        INSERT INTO processed.margin_backlog
                        (channel, invoice, iccid, sku, product_name,
                         product_type, order_date, harga_jual, reason, resolved)
                        VALUES (:channel, :invoice, :iccid, :sku, :product_name,
                                :product_type, :order_date, :harga_jual, :reason, :resolved)
                    """), row.to_dict())
        print(f"  Backlog DB: {len(backlog_db)} baris disimpan")

    print(f"  Excluded: {len(no_match)} baris (masuk backlog)")

    print("\nExtracting fallback candidates dari diagnostics...")
    fallback_cands = extract_fallback_candidates(diagnostics_df)
    if not fallback_cands.empty:
        print(f"  Kandidat fallback ({len(fallback_cands)} pairs):")
        print(fallback_cands.to_string(index=False))
        cand_path = OUTPUT_DIR / "sku_fallback_candidates.json"
        top = (
            fallback_cands[~fallback_cands["Already_In_Fallback"]]
            .sort_values("Count_ICCID", ascending=False)
            .drop_duplicates("Order_Country")
        )
        with open(cand_path, "w") as f:
            import json as _json
            _json.dump(dict(zip(top["Order_Country"], top["DB_Country"])), f, indent=4)
        print(f"  Kandidat baru (belum di fallback) → {cand_path}")
    else:
        print("  Tidak ada country mismatch baru ditemukan")

    # cek backlog lama yang sudah bisa di-resolve
    with engine.connect() as conn:
        old_backlog = pd.read_sql(text("""
            SELECT id, invoice, iccid, sku
            FROM processed.margin_backlog
            WHERE resolved = FALSE AND reason IN ('ACTIVE', 'NO_ORDER_DATA')
        """), conn)

    if not old_backlog.empty:
        from processors.behaviour_factor import build_sku
        final_sku = final.copy()
        final_sku["sku_match"] = final_sku["package"].apply(lambda p: build_sku(p).upper())
        resolvable = old_backlog.merge(
            final_sku[["iccid", "sku_match", "real_cost_idr"]],
            left_on=["iccid", "sku"], right_on=["iccid", "sku_match"], how="inner"
        )
        resolvable = resolvable[resolvable["real_cost_idr"].notna()]
        if not resolvable.empty:
            with engine.begin() as conn:
                for bid in resolvable["id"].tolist():
                    conn.execute(text("""
                        UPDATE processed.margin_backlog SET resolved = TRUE WHERE id = :id
                    """), {"id": bid})
            print(f"  Backlog resolved: {len(resolvable)} baris sudah bisa dihitung")

    # ── NEGARA DETAIL ─────────────────────────────────────────
    country_usage["Rate_CNY"]     = country_usage["country"].apply(get_rate)
    country_usage["Cost_CNY"]     = country_usage["usage_mb"] / 1024 * country_usage["Rate_CNY"]
    iccid_total                   = country_usage.groupby("iccid")["usage_mb"].transform("sum")
    country_usage["usage_pct"]    = (country_usage["usage_mb"] / iccid_total * 100).round(1)
    country_usage["country_name"] = country_usage["country"].apply(
        lambda x: MCC_MAP.get(str(extract_mcc(x)), x)
    )
    country_detail = (
        country_usage
        .sort_values(["iccid", "usage_pct"], ascending=[True, False])
        .groupby("iccid")
        .apply(lambda d: ", ".join(
            f"{row['country_name']} ({row['usage_pct']}%, {row['Cost_CNY']:.2f} CNY)"
            for _, row in d.iterrows()
        ), include_groups=False)
        .reset_index()
    )
    country_detail.columns = ["iccid", "Negara_Detail"]

    # ── BUILD REPORT ──────────────────────────────────────────
    df = orders.copy()
    df = df.rename(columns={
        "real_cost_cny":  "Real_Cost_CNY",
        "real_cost_idr":  "Real_Cost_IDR",
        "total_usage_mb": "Total_Usage_MB",
        "total_quota_mb": "Total_Quota_MB",
    })
    df["Real_Cost_CNY"] = df["Real_Cost_CNY"].fillna(0).round(2)
    df["Real_Cost_IDR"] = (df["Real_Cost_CNY"] * CNY_RATE).round(2)
    df["Total_Usage_Display"] = df["Total_Usage_MB"].apply(
        lambda x: f"{round(x/1024, 2)} GB" if pd.notna(x) and x >= 1024
        else f"{round(x, 2)} MB" if pd.notna(x) else "-"
    )
    df["Total_Quota_Display"] = df["Total_Quota_MB"].apply(
        lambda x: f"{round(x/1024, 2)} GB" if pd.notna(x) and x >= 1024
        else f"{round(x, 2)} MB" if pd.notna(x) else "-"
    )
    df["Ratio"] = (
        df["Total_Usage_MB"].fillna(0) / df["Total_Quota_MB"].replace(0, np.nan)
    ).fillna(0).round(4)
    df = df.merge(country_detail.rename(columns={"iccid": "ICCID"}), on="ICCID", how="left")
    df["Negara_Detail"] = df["Negara_Detail"].fillna("UNKNOWN")

    df["Komisi_Pct"] = df.apply(lambda r: get_commission(r["Channel"], r["Product_Type"]), axis=1)
    df["Komisi_IDR"]  = (df["Harga_Jual"] * df["Komisi_Pct"]).round(2)
    df["Net_Revenue"] = (df["Harga_Jual"] - df["Komisi_IDR"]).round(2)
    df["Margin_IDR"]  = (df["Net_Revenue"] - df["Real_Cost_IDR"]).round(2)
    df["Margin_Pct"]  = np.where(
        df["Net_Revenue"] > 0,
        (df["Margin_IDR"] / df["Net_Revenue"]).round(4), 0
    )
    df["Status"] = np.select(
        [df["Margin_Pct"] < 0, df["Margin_Pct"] >= 0.30],
        ["RUGI", "BAGUS"], default="NORMAL"
    )

    # ── APPLY DATE FILTER (output only) ───────────────────────
    df_all = df.copy()  # keep full data for total context
    if args.start_date or args.end_date:
        print(f"\nApplying date filter for output...")
        df = apply_date_filter(df, args.start_date, args.end_date)
        backlog = apply_date_filter(backlog, args.start_date, args.end_date)

    # ── SUMMARIES ─────────────────────────────────────────────
    summary_sku = (
        df.groupby("SKU")
        .agg(
            Total_Invoice     = ("Invoice",       "nunique"),
            Total_ICCID       = ("ICCID",         "count"),
            Total_Pendapatan  = ("Harga_Jual",    "sum"),
            Total_Komisi      = ("Komisi_IDR",    "sum"),
            Total_Net_Revenue = ("Net_Revenue",   "sum"),
            Total_Cost_IDR    = ("Real_Cost_IDR", "sum"),
            Total_Margin_IDR  = ("Margin_IDR",    "sum"),
        )
        .reset_index()
    )
    for col in ["Total_Pendapatan", "Total_Komisi", "Total_Net_Revenue",
                "Total_Cost_IDR", "Total_Margin_IDR"]:
        summary_sku[col] = summary_sku[col].round(2)
    summary_sku["Avg_Margin_Pct"] = (
        summary_sku["Total_Margin_IDR"] / summary_sku["Total_Net_Revenue"].replace(0, np.nan)
    ).round(4)
    summary_sku["Status"] = np.select(
        [summary_sku["Avg_Margin_Pct"] < 0, summary_sku["Avg_Margin_Pct"] >= 0.30],
        ["RUGI", "BAGUS"], default="NORMAL"
    )
    summary_sku = summary_sku.sort_values("Avg_Margin_Pct")

    summary_month = (
        df.groupby("Month")
        .agg(
            Total_Invoice          = ("Invoice",       "nunique"),
            Total_ICCID            = ("ICCID",         "count"),
            Total_Pendapatan_Gross = ("Harga_Jual",    "sum"),
            Total_Komisi           = ("Komisi_IDR",    "sum"),
            Total_Net_Revenue      = ("Net_Revenue",   "sum"),
            Total_Cost_IDR         = ("Real_Cost_IDR", "sum"),
            Total_Margin_IDR       = ("Margin_IDR",    "sum"),
            ICCID_Rugi             = ("Status",        lambda x: (x == "RUGI").sum()),
            ICCID_Normal           = ("Status",        lambda x: (x == "NORMAL").sum()),
            ICCID_Bagus            = ("Status",        lambda x: (x == "BAGUS").sum()),
        )
        .reset_index()
    )
    for col in ["Total_Pendapatan_Gross", "Total_Komisi", "Total_Net_Revenue",
                "Total_Cost_IDR", "Total_Margin_IDR"]:
        summary_month[col] = summary_month[col].round(2)
    summary_month["Avg_Margin_Pct"] = (
        summary_month["Total_Margin_IDR"] / summary_month["Total_Net_Revenue"].replace(0, np.nan)
    ).round(4)
    summary_month = summary_month.sort_values("Month")

    total_margin_idr  = round(df["Margin_IDR"].sum(), 2)
    total_net_revenue = round(df["Net_Revenue"].sum(), 2)
    total = pd.DataFrame([{
        "Total_Invoice":          df["Invoice"].nunique(),
        "Total_ICCID":            len(df),
        "Total_Pendapatan_Gross": round(df["Harga_Jual"].sum(), 2),
        "Total_Komisi":           round(df["Komisi_IDR"].sum(), 2),
        "Total_Net_Revenue":      total_net_revenue,
        "Total_Cost_IDR":         round(df["Real_Cost_IDR"].sum(), 2),
        "Total_Margin_IDR":       total_margin_idr,
        "Avg_Margin_Pct":         round(total_margin_idr / total_net_revenue, 4) if total_net_revenue else 0,
        "ICCID_Rugi":             (df["Status"] == "RUGI").sum(),
        "ICCID_Normal":           (df["Status"] == "NORMAL").sum(),
        "ICCID_Bagus":            (df["Status"] == "BAGUS").sum(),
        "Total_Backlog":          len(backlog),
    }])

    # add date range info
    if args.start_date or args.end_date:
        total["Date_Range"] = f"{args.start_date or 'awal'} s/d {args.end_date or 'akhir'}"

    print("\nSUMMARY TOTAL:")
    print(total.to_string(index=False))

    # ── SAVE TO EXCEL ─────────────────────────────────────────
    print(f"\nSaving to {OUTPUT}...")

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        if "Match_Method" not in df.columns:
            df["Match_Method"] = "SKU_MATCH"

        summary_month.to_excel( writer, sheet_name="Summary per Bulan", index=False)
        total.to_excel(         writer, sheet_name="Summary Total",      index=False)
        backlog.to_excel(       writer, sheet_name="Backlog",            index=False)

        # diagnostic sheet — untuk review NO_ORDER_DATA
        if not diagnostics_df.empty:
            diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)

        # fallback candidates sheet
        if not fallback_cands.empty:
            fallback_cands.to_excel(writer, sheet_name="Fallback Candidates", index=False)

        wb = writer.book

        # Settings sheet — referenced by Real_Cost_IDR formula in Detail sheet
        _write_settings_sheet(wb, CNY_RATE)

        # Detail sheet — Excel formulas so client can verify calculations
        ws_detail = wb.create_sheet("Detail per ICCID", 0)
        _write_detail_formula_sheet(ws_detail, df)

        # Summary per SKU — SUMIF formulas referencing Detail sheet
        ws_sku = wb.create_sheet("Summary per SKU", 1)
        _write_summary_sku_formula_sheet(ws_sku, df, len(df) + 1)

        IDR_COLS = ["Harga_Jual", "Komisi_IDR", "Net_Revenue", "Real_Cost_IDR",
                    "Margin_IDR", "Total_Pendapatan", "Total_Komisi", "Total_Net_Revenue",
                    "Total_Cost_IDR", "Total_Margin_IDR", "Total_Pendapatan_Gross"]
        PCT_COLS = ["Margin_Pct", "Avg_Margin_Pct"]

        style_sheet(wb["Summary per Bulan"], pct_cols=PCT_COLS, idr_cols=IDR_COLS)
        style_sheet(wb["Summary Total"],     pct_cols=PCT_COLS, idr_cols=IDR_COLS)

        if not diagnostics_df.empty:
            ws_diag = wb["Diagnostics"]
            FILL_HEADER = PatternFill("solid", fgColor="2C3E50")
            FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
            FILL_WARN   = PatternFill("solid", fgColor="FFDDDD")
            for cell in ws_diag[1]:
                cell.fill = FILL_HEADER; cell.font = FONT_HEADER
            for row in ws_diag.iter_rows(min_row=2):
                for cell in row: cell.fill = FILL_WARN
            for col in ws_diag.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws_diag.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        if not fallback_cands.empty:
            ws_fc = wb["Fallback Candidates"]
            FILL_HEADER  = PatternFill("solid", fgColor="2C3E50")
            FONT_HEADER  = Font(color="FFFFFF", bold=True, size=11)
            FILL_NEW     = PatternFill("solid", fgColor="FFF0B3")   # kuning = belum di fallback
            FILL_EXIST   = PatternFill("solid", fgColor="CCFFCC")   # hijau  = sudah ada
            headers      = {cell.value: cell.column for cell in ws_fc[1]}
            already_col  = headers.get("Already_In_Fallback")
            for cell in ws_fc[1]:
                cell.fill = FILL_HEADER; cell.font = FONT_HEADER
            for row in ws_fc.iter_rows(min_row=2):
                is_exist = already_col and str(ws_fc.cell(row[0].row, already_col).value).lower() == "true"
                fill = FILL_EXIST if is_exist else FILL_NEW
                for cell in row: cell.fill = fill
            for col in ws_fc.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws_fc.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    print(f"\nMargin report selesai! → {OUTPUT}")
    if args.start_date or args.end_date:
        print(f"  Date range: {args.start_date or 'awal'} s/d {args.end_date or 'akhir'}")
    if not diagnostics_df.empty:
        print(f"  Diagnostics: {len(diagnostics_df)} ICCID perlu di-review (lihat sheet Diagnostics)")
    if not fallback_cands.empty:
        new_count = (~fallback_cands["Already_In_Fallback"]).sum()
        print(f"  Fallback Candidates: {len(fallback_cands)} pairs "
              f"({new_count} baru) → lihat sheet + {OUTPUT_DIR}/sku_fallback_candidates.json")